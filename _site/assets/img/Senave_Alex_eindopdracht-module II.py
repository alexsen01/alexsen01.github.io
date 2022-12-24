import datetime
import random
from openpyxl import load_workbook,Workbook #afzonderlijk geladen anders foutmelding bij creëren txt file, ipv #from  openpyxl import *
from openpyxl.chart import BarChart3D, Reference #https://openpyxl.readthedocs.io/en/stable/charts/bar.html

class MyProducten():
    def __init__(self):
        #begincombinatie is: product,prijs,voorraad
        self.dranken=["thee",12,1,"koffie",15,2,"melk",20,3,"cola",30,4,"wijn",35,6]
        self.etenswaren=["varkensvlees",1,25,"biefstuk",20,30,"cordon blue",20,40,"bbq vlees",10,50]
        self.verzorging=["nivea creme",15,3,"scheermesjes",20,10]
        self.bureau=["tippex",5,4,"bic",1,25,"potlood",1,15,"gom",2,8]
        self.categorieën={"dranken":self.dranken,"etenswaren":self.etenswaren,"verzorging":self.verzorging,"bureau":self.bureau}

class MyKlant():
    def __init__(self):
        self.naam=input('wat is uw familienaam?: ') 
        self.voornaam=input('wat is uw voornaam?: ')
        self.stad=input('wat is uw stad?: ') 
        self.straat=input('wat is uw straat?: ') 
        self.huis=input('wat is uw huisnr?: ')  
        self.gsm=input('wat is uw gsm nr?: ')  

class MyNieweKlant(MyKlant):
    def __init__(self):
        super().__init__()
    def nieuwe_klant(self):
        try:
            wb = load_workbook('Klanten.xlsx')
            ws = wb["Klanten"]
            self.klantnummer=ws.max_row
        except:
            wb = Workbook() 
            del wb["Sheet"]
            ws=wb.create_sheet()
            ws.title = "Klanten"
            ws['A1'].value="klantnr"
            ws['B1'].value="naam"
            ws['C1'].value="voornaam"
            ws['D1'].value="stad"
            ws['E1'].value="straat"
            ws['F1'].value="huisnr"
            ws['G1'].value="gsm"
            ws['H1'].value="datum"
            self.klantnummer=1
        ws.cell(row=self.klantnummer+1,column=1).value=self.klantnummer
        ws.cell(row=self.klantnummer+1,column=2).value=self.naam
        ws.cell(row=self.klantnummer+1,column=3).value=self.voornaam
        ws.cell(row=self.klantnummer+1,column=4).value=self.stad
        ws.cell(row=self.klantnummer+1,column=5).value=self.straat
        ws.cell(row=self.klantnummer+1,column=6).value=self.huis
        ws.cell(row=self.klantnummer+1,column=7).value=self.gsm
        ws.cell(row=self.klantnummer+1,column=8).value=datetime.date.today()
        wb.save('Klanten.xlsx')
        wb.close()
        return self.klantnummer

def producteninladen():
    print('producten inladen')
    rijteller=0 
    cat_teller=0
    for categorie in MyProducten().categorieën.keys():
        try:
            wb = load_workbook('Winkel.xlsx')
            ws = wb[categorie]
        except:
            #creëren van de workbook (dit mag enkel maar bij eerste loop, want bij tweede categorie bestaat die al)
            #controle dmv de rijteller: enkel bij begin van eerste loop is deze=0 
            if rijteller==0:
                wb = Workbook()
                del wb["Sheet"]
            else:
                wb = load_workbook('Winkel.xlsx')
            ws=wb.create_sheet(categorie)
            ws['A1'].value="Categorie"
            ws['B1'].value="Product"
            ws['C1'].value="Voorraad"
            ws['D1'].value="Prijs"
            ws['E1'].value="Promotiekorting"
            for rijteller in range(0,len(MyProducten().categorieën[categorie]),3): #rijteller wordt hier geupdated
                rij=int(rijteller/3+2)
                ws.cell(row=rij,column=1).value=categorie
                product=MyProducten().categorieën[categorie][rijteller]
                ws.cell(row=rij,column=2).value=product
                aantal=MyProducten().categorieën[categorie][rijteller+1]
                ws.cell(row=rij,column=3).value=aantal
                prijs=MyProducten().categorieën[categorie][rijteller+2]
                ws.cell(row=rij,column=4).value=prijs
            wb.save('Winkel.xlsx')
        cat_teller+=1
        wb.close()
    
def bijbestellingen_prijsaanpassingen_promotiekortingen(actie):
    try:
        wb = load_workbook('Winkel.xlsx')
    except:
        producteninladen()
        wb = load_workbook('Winkel.xlsx')
    for i in wb.sheetnames:
        check1=0
        ws=wb[i]
        print('\ncategorie is: ',ws.title) 
        while check1==0:
            try:
                if actie==3:
                    vraag=input('bijbestellingen voor deze categorie? (j/n): ')
                elif actie==4:
                    vraag=input('prijsaanpassingen voor deze categorie? (j/n): ')
                elif actie==5:
                    vraag=input('promotiekortingen voor deze categorie? (j/n): ')
                if vraag not in ("j","n"):
                    print('verkeerd antwoord')
                else:
                    check1=1
            except:
                print('verkeerde input')
            if vraag=="j":
                teller=0
                if actie==3:
                    print('dit zijn al de producten uit deze categorie en de voorraad ervan:')
                    for teller in range(1,ws.max_row): 
                        print(teller,': ',ws.cell(row=teller+1,column=2).value,' : ',ws.cell(row=teller+1,column=3).value)
                        teller +=1
                elif actie==4:
                    print('dit zijn al de producten uit deze categorie en de huidige prijs ervan:')
                    for teller in range(1,ws.max_row):
                        print(teller,': ',ws.cell(row=teller+1,column=2).value,' : ',ws.cell(row=teller+1,column=4).value)
                        teller +=1                        
                else:
                    print('dit zijn al de producten uit deze categorie:')
                    for teller in range(1,ws.max_row):
                        print(teller,': ',ws.cell(row=teller+1,column=2).value)
                        teller +=1                        
                check2=0
                while check2==0:
                    try:
                        if actie==3:
                            vraag2=int(input('voor welk product bijbestellen? (nummer): '))
                        elif actie==4:
                            vraag2=int(input('voor welk product prijsaanpassing? (nummer): '))
                        elif actie==5:
                            vraag2=int(input('voor welk product promotiekorting? (nummer): '))
                        if not (1<=vraag2<=ws.max_row-1):
                            print('verkeerd nummer')
                        else:
                            check2=1
                    except:
                        print('verkeerde input')
                    else:
                        if check2==1:
                            check3=0
                            rij=vraag2+1
                            while check3==0:
                                try:
                                    if actie==3:
                                        bijbestellen3=int(input('hoeveel wilt u er bijbestellen?: '))
                                        if bijbestellen3<0:
                                            print('verkeerd aantal: moet >=0 zijn')
                                        else:
                                            nieuw3=bijbestellen3+ws.cell(row=rij,column=3).value
                                            check3=1
                                    elif actie==4:
                                        aanpassing3=int(input('wat is de nieuwe prijs voor dit product?: '))
                                        if aanpassing3<0:
                                            print('verkeerde prijs: moet >=0 zijn')
                                        else:
                                            nieuw3=aanpassing3
                                            check3=1
                                    elif actie==5:
                                        korting3=int(input('hoeveel procent korting voor dit product?: '))
                                        if not (0<= korting3 <=100):
                                            print('verkeerd percentage: moet 0<= % <=100 zijn')
                                        else:
                                            nieuw3=korting3
                                            check3=1
                                except:
                                    print('verkeerde input')
                                else:
                                    if check3==1:
                                        ws.cell(row=rij,column=actie).value=nieuw3 # de waarde voor actie komt overeen met de kolom die moet veranderd worden
                                        wb.save('Winkel.xlsx')
                            check4=0
                            while check4==0:
                                try:
                                    vraag4=input('nog andere producten van deze categorie? (j/n): ') 
                                    if vraag4 not in ('j','n'):
                                        print('verkeerde input')
                                    else:
                                        check4=1
                                except:
                                    print('verkeerde input')
                            if vraag4=="j":
                                check2=0
    wb.close()

def sub_sub_nieuwe_producten(prod,ws,wb): #het effectief toevoegen van nieuwe producten aan de xls sheet
    check=0
    while check==0:
        try:
            voorraad=int(input('hoeveel van deze producten?: '))
            prijs=int(input('wat is de prijs van deze?: '))
            prom=input('is er een promotiekorting voor dit product? (j/n): ')
            if prom=="j":
                promotiekorting=int(input('hoeveel % korting?: '))
            elif prom=="n":
                promotiekorting=0                
            if ((0<voorraad) and (0<prijs) and (prom in ('j','n'))):
                check=1
            else:
                print('verkeerde input-1')
        except:
            print('verkeerde input-2')
        else:
            if check==1:
                rij=ws.max_row+1
                ws.cell(row=rij,column=1).value=ws.title
                ws.cell(row=rij,column=2).value=prod
                ws.cell(row=rij,column=3).value=voorraad
                ws.cell(row=rij,column=4).value=prijs
                ws.cell(row=rij,column=5).value=promotiekorting
                wb.save('Winkel.xlsx')

def sub_nieuwe_of_verwijderen_producten(wb,ws,vraag): #keuzemenu producten toevoegen/verwijderen (ook volledige categorie)
    check2=0
    while check2==0:
        if vraag==2:
            print('dit zijn de huidige producten in deze categorie:')
            for teller in range(1,ws.max_row):
                print(teller,': ',ws.cell(row=teller+1,column=2).value)
                teller +=1
        try:
            if vraag==1:
                prod=input('welk product wilt u toevoegen? (0 om te stoppen): ')
                check2=1
                if prod=="0":
                    vraag=3
            elif vraag==2:
                keuze=int(input('welk product verwijderen? (0 om te stoppen, -1 voor de volledige categorie): '))
                if (-1<=keuze<ws.max_row): #mag =0 zijn, om uit de loop te kunnen gaan
                    check2=1
                else:
                    print ('verkeerde keuze')
                if keuze==0:
                    vraag=3
                elif keuze==-1: 
                    vraag=4
        except:
            print('verkeerde input') 
        else:
            if vraag==1:
                for teller in range(1,ws.max_row):
                    if prod==ws.cell(row=teller+1,column=2).value: 
                        print('dit product bestaat reeds in deze categorie')
                        check2=0
                    teller +=1
                if check2==1:
                    sub_sub_nieuwe_producten(prod,ws,wb) #naar de effectieve toevoeging in de xls sheet 
            elif vraag==2:
                ws.delete_rows(keuze+1) #het effectief verwijderen uit de xls sheet
            elif vraag==4:
                titel=ws.title
                del wb[titel] 
                loop1=0
            if vraag in (1,2): 
                check3=0
                while check3==0:
                    try:
                        if vraag==1:
                            nog=input('nog andere producten toevoegen aan deze categorie? (j/n): ')
                        elif vraag==2:
                            nog=input('nog andere producten verwijderen uit deze categorie? (j/n): ')
                        if nog in ('j','n'):
                            check3=1
                        else:
                            print('verkeerde input')
                    except:
                        print('verkeerde input')
                    else:
                        if nog=='j':
                            check2=0
                        elif nog=='n':
                            check4=0
                            while check4==0:
                                try:
                                    if vraag==1:
                                        andere_cat=input('andere producten toevoegen aan andere categorie? (j/n): ')
                                    elif vraag==2:
                                        andere_cat=input('andere producten verwijderen uit andere categorie? (j/n): ')
                                    if andere_cat in ('j','n'):
                                        check4=1
                                    else:
                                        print('verkeerde input')
                                except:
                                    print('verkeerde input')
                                else:
                                    if andere_cat=='j':
                                        loop1=1
                                    elif andere_cat=='n':
                                        loop1=0
            elif vraag==3:
                loop1=0
    return loop1

def nieuwe_of_verwijderen_producten(): #keuzemenu categorie toevoegen/verwijderen
    try:
        wb = load_workbook('Winkel.xlsx')
    except:
        producteninladen()
        wb = load_workbook('Winkel.xlsx')
    check=0
    while check==0:
        try:
            vraag=int(input('(1): wilt u nieuwe producten toevoegen of (2) bestaande producten verwijderen?: '))
            if vraag in (1,2):
                check=1
            else:
                print('verkeerde input')
        except:
            print('verkeerde input')
    loop=1
    while loop==1:
        print('\ndit zijn de bestaande categorieën:')
        teller=1
        for i in wb.sheetnames:
            print('\ncategorie',teller,'is:',i)
            teller+=1
        if vraag==1:
            check1=0
            while check1==0:
                try:
                    bestaand=input('\néén van deze categorieën? (j/n): ')
                    if bestaand in ('j','n'):
                        check1=1
                    else:
                        print('verkeerde input')
                except:
                    print('verkeerde input')
        elif vraag==2:
            bestaand="j" #voor volgende if/elif
        if bestaand=='j':
            aantal=teller-1
            check1=0
            while check1==0:
                try:
                    keuze_cat=int(input('welke categorie?: 1-'+str(aantal)+': '))
                    if keuze_cat in range(1,aantal+1): 
                        check1=1
                    else:
                        print('verkeerde keuze-1')
                except:
                    print('verkeerde keuze-2')
            ws=wb[wb.sheetnames[keuze_cat-1]]
        elif bestaand=='n': 
            check1=0
            while check1==0:
                try:
                    nieuwe_cat=input('wat is de naam voor de nieuwe categorie?: ')
                    check1=1
                except:
                    print('verkeerde input')
            ws=wb.create_sheet(nieuwe_cat)
            ws['A1'].value="Categorie"
            ws['B1'].value="Product"
            ws['C1'].value="Voorraad"
            ws['D1'].value="Prijs"
            ws['E1'].value="Promotiekorting"
        loop=sub_nieuwe_of_verwijderen_producten(wb,ws,vraag) #effectief invoeren / verwijderen van producten #     
    wb.save('Winkel.xlsx')
    wb.close()

def sub_klant(): #hier wordt klant gechecked en eventueel in xls sheet gezet
    check=0
    while check==0:
        try:
            klant_jn=input('bent u reeds klant? (j/n): ')
            if klant_jn in ("j","n"):
                check=1
        except:
            print('verkeerde input')
        else:
            if klant_jn=="j":  #reeds klant
                try:
                    wb = load_workbook('Klanten.xlsx')
                    ws = wb["Klanten"]
                except:
                    print('klantenbestand bestaat nog niet')
                    check=0 #om terug de hoofdloop uit te voeren
                else:
                    klantnr=0
                    while klantnr==0:
                        try:
                            klantnr=int(input('wat is uw klantnummer?: '))
                            if not (0<klantnr<ws.max_row): 
                                print('dit klantnummer is niet gekend')
                                klantnr=1
                                check=0
                        except:
                            print('verkeerde input')
                            klantnr=0
                    wb.close()
                klant_jn="n" 
            elif klant_jn=="n":  #nog geen klant
                check2=0
                while check2==0:
                    try:
                        klant_jn2=input('wilt u klant worden? (5% extra korting) (j/n): ')
                        if klant_jn2 in ("j","n"):
                            check2=1
                    except:
                        print('verkeerde input')
                    else: 
                        if klant_jn2=="j":  #nieuwe klant invoeren
                            r=MyNieweKlant()
                            klantnr=r.nieuwe_klant()
                        elif klant_jn2=="n": #niet als nieuwe klant => klantnr=0 stellen
                            klantnr=0
    return klantnr
   
def sub_sub_aankopen(klantnr,ws,wb,cat,prod_code,product,aantal,prijs,korting,nieuw):  #de aankopen registreren
    check=0
    titel=str(datetime.date.today())
    try:
        wb1 = load_workbook('Inkopen.xlsx') 
    except:
        wb1 = Workbook() #wb1 bestond nog niet, nu creëren
        del wb1["Sheet"]
        ws1=wb1.create_sheet() #nieuwe worksheet
        ws1.title=titel
        header=1
    else:
        for dag in wb1.sheetnames:
            if dag==titel:
                ws1 = wb1[titel] #worksheet van vandaag bestond reeds, deze wordt nu ws1
                header=0
                if nieuw==1: #als compleet nieuw order, 
                    aankoopnr=ws1.cell(row=ws1.max_row,column=1).value+1 #dan nieuw aankoopnr=oud+1,
                else:
                    aankoopnr=ws1.cell(row=ws1.max_row,column=1).value #anders behouden we het oude aankoopnr
                check=1 #voor de volgende stap over te slaan
        if check==0: #vorige loop vond geen ws'n voor vandaag
            ws1=wb1.create_sheet() 
            ws1.title=titel
            header=1        
    if header==1:
        ws1['A1'].value="aankoopnr"
        ws1['B1'].value="klantnr"
        ws1['C1'].value="cat"            
        ws1['D1'].value="product"
        ws1['E1'].value="aantal"
        ws1['F1'].value="prijs"
        ws1['G1'].value="aantalprijs"
        ws1['H1'].value="promotiekorting"
        ws1['I1'].value="eindprijs" 
        aankoopnr=1
    rij=int(ws1.max_row+1)
    ws1.cell(row=rij,column=1).value=aankoopnr
    ws1.cell(row=rij,column=2).value=klantnr
    ws1.cell(row=rij,column=3).value=cat
    ws1.cell(row=rij,column=4).value=product
    ws1.cell(row=rij,column=5).value=aantal
    ws1.cell(row=rij,column=6).value=prijs
    ws1.cell(row=rij,column=7).value=prijs*aantal
    ws1.cell(row=rij,column=8).value=korting
    ws1.cell(row=rij,column=9).value=prijs*aantal*(1-korting/100)
    wb1.save('Inkopen.xlsx')
    wb1.close()
    ws.cell(row=prod_code+1,column=3).value -=aantal #voorraad wordt verminderd met het aantal aangekocht
    wb.save('Winkel.xlsx')
    return aankoopnr

def sub_rekening(aankoopnr,klantnr):
    titel=str(datetime.date.today())
    som=0
    if klantnr != 0: #als klantnummer=0 is, dan is het geen geregistreerde klant
        klantkorting=0.05
    else:
        klantkorting=0
    try:
        wb1 = load_workbook('Inkopen.xlsx')
        ws1=wb1[titel]
    except:
        print('fout bij openen Inkopen bestand')
    else:
        klantrekening=titel+"_klantnr"+str(klantnr)+"_aankoopnr"+str(aankoopnr)+".txt"
        bestand=open(klantrekening,"w") 
        a="Winkel: Python2\n\ndatum: "+titel+"  tijdstip: "+str(datetime.datetime.now().strftime("%H:%M:%S"))+"\n\nklantnummer: "+str(klantnr)
        bestand.write(a+"\n"+"\n")
        b="afzonderlijke aankopen:"
        print("\n"+b)
        bestand.write(b)
        for rij in range(1,ws1.max_row+1):
            if ws1.cell(row=rij,column=1).value==aankoopnr:
                product=str(ws1.cell(row=rij,column=4).value) 
                aantal=str(ws1.cell(row=rij,column=5).value)
                stukprijs=str(format(round(ws1.cell(row=rij,column=6).value,2),'.2f'))
                aantalprijs=str(format(round(ws1.cell(row=rij,column=7).value,2),'.2f'))
                korting=ws1.cell(row=rij,column=7).value*(ws1.cell(row=rij,column=8).value/100)
                eindprijs=ws1.cell(row=rij,column=9).value
                c1="-product: "+product
                c2="--stukprijs: "+stukprijs+" aantal: "+aantal+" aantalprijs: "+aantalprijs+" korting: "+str(format(round(korting,2),'.2f'))+" eindprijs: "+str(format(round(eindprijs,2),'.2f'))
                print(c1)
                print(c2)
                bestand.write("\n"+c1)
                bestand.write("\n"+c2)
                som +=eindprijs
        c1="aankoopbedrag: "+str(format(round(som,2),'.2f'))
        bestand.write("\n"+c1)
        print(c1)
        klantkortingbedrag=som*klantkorting
        if klantkorting==0.05:
            c2="\nu heeft recht op een klantkorting van "+str(format(klantkorting*100,'.0f'))+"%: "+str(format(round(klantkortingbedrag,2),'.2f'))
        else:
            c2="\nenkel een klantenkorting als u klant bent"
        print(c2)
        bestand.write("\n"+c2)
        eindsom=som-klantkortingbedrag
        c3="tussenbedrag: "+str(format(round(eindsom,2),'.2f'))
        bestand.write("\n"+c3)
        print(c3)
        bedragkorting=0
        if eindsom>=200:
            bedragkorting=15
        elif eindsom>=150:
            bedragkorting=10
        elif eindsom>=100:
            bedragkorting=5
        if eindsom>=100:
            c4="korting door bedraggrootte: "+str(format(bedragkorting,'.2f'))
        else:
            c4="enkel een bedragkorting bij een tussenbedrag van minimaal 100 euro"
        print(c4)
        bestand.write("\n"+c4)
        eindeindbedrag=round(eindsom-bedragkorting,2)
        d="\neindbedrag te betalen: "+str(format(eindeindbedrag,'.2f'))
        bestand.write("\n"+d)
        print(d)
        bestand.close()
        sub_betaling(eindeindbedrag) 

def sub_sub_cashbetaling(eindeindbedrag):
    coupures=[500,200,100,50,20,10,5,2,1,0.2,0.1,0.05,0.02]
    totaal=0
    while totaal<eindeindbedrag: #om op het einde van de coupures, de vragen opnieuw te beginnen als nog niet genoeg betaald
        check0=0
        for coup in coupures:
            if check0==0:
                check=0
                while check==0:
                    try:
                        aantal=input('hoeveel van '+str(coup)+": (geef a in om te herbeginnen)")
                        if aantal=="a":
                            totaal=0
                            check0=1
                            break
                        elif not (0<=int(aantal)):
                            print('verkeerd')
                        else:
                            check=1
                    except:
                        print('verkeerd')
                    else:
                        if aantal!="a": 
                            aantal=int(aantal)
                            totaal+=(coup*aantal)
                            rest=round(eindeindbedrag-totaal,2)
                            if totaal<eindeindbedrag:
                                print('nog te betalen:',format(rest,'.2f'))
                                check0=0
                            else:
                                print('u krijgt volgend bedrag terug:',format(-rest,'.2f'))
                                check=1
                                check0=1
                        elif aantal=="a":
                            check0=1
                            check=1
        if (totaal<eindeindbedrag) and (aantal!="a"):
            print('te weinig betaald, betaal opnieuw')
   
def sub_betaling(eindeindbedrag):
    check=0
    print('hoe doet u de betaling?')
    check1=0
    bet1=0
    bet2=0
    while check==0:
        try:
            if check1==0:
                bet1=int(input('1: bankkaart 2: cash 3: maaltijdcheques  \n?:'))
            else:
                bet2=int(input('1: bankkaart 2: cash \n?:')) #voor de 2e iteratie hierover (betalen van het restbedrag na maaltijdcheques=onvoldoende saldo)
            if bet1 in (1,2,3) or bet2 in (1,2):
                check=1
            else:
                print('verkeerde input')
        except:
            print('verkeerde input')
        else:
            if bet1==1:
                print('u heeft betaald met bankkaart')
            elif bet2==1:
                print('u heeft de rest betaald met bankkaart')
            elif bet1==2 or bet2==2:
                sub_sub_cashbetaling(eindeindbedrag)
            elif bet1==3:
                cheq=random.randint(0,int(100*eindeindbedrag))/100
                print('het bedrag dat betaald wordt met maaltijdcheques is: ',format(cheq,'.2f'))
                eindeindbedrag -=cheq
                print('hoe wilt u het resterende bedrag betalen?: ',format(eindeindbedrag,'.2f'))
                check=0
                check1=1

def sub_aankopen(wb,klantnr): #hier kan de klant kiezen uit de lijst met producten/aantal
    for categorie in wb.sheetnames:
        ws=wb[categorie]       
        print('\nvoor ',categorie,' is er de mogelijkheid uit:') #loop over alle mogelijkheden, met cijfer
        for teller in range(1,ws.max_row):
            print(teller,': ',ws.cell(row=teller+1,column=2).value)
            teller +=1
    nog_prod="j"
    teller=1
    nieuw=1#parameter om in sub_sub_aankopen in eerste kolom aan te duiden dat het een compleet nieuw order is
    print('\n')
    for i in wb.sheetnames: #dit xlsx bestand bevat al de categorieën/producten/...
        print('categorie',teller,': ',i)
        teller+=1
    aankoopnr=0
    while nog_prod=="j":
        sel=0
        while sel==0:
            try:
                cat=int(input('\nwelke categorie kiest u?: (1-'+str(teller-1)+'): '))
                if (cat <1) or (cat > teller-1):
                    print('verkeerde keuze van categorie')
                    sel=0
                else:
                    prod_code=int(input('\nwelk product kiest u? (zie bovenstaande lijst): '))
                    ws=wb[wb.sheetnames[cat-1]] #wb.sheetnames geeft een [], dus hier kan uit gekozen worden voor naam van de worksheet
                    if (prod_code<1) or (prod_code > ws.max_row-1): 
                        print('verkeerde keuze van product')
                        sel=0
                    else:
                        maximum=ws.cell(row=prod_code+1,column=3).value
                        while sel==0:
                            try:
                                aantal=int(input(str(ws.cell(row=prod_code+1,column=2).value)+': hoeveel kiest u er hiervan? Maximum is '+str(maximum)+': '))
                                if aantal > maximum:
                                    print('verkeerd aantal, max is:',maximum)
                                elif aantal<0:
                                    print('minder dan 0 gekozen')
                                elif aantal==0:
                                    print('0 gekozen: dit product wordt uit de shopping lijst gehaald')
                                    break
                                else:
                                    sel=1
                            except:
                                print('verkeerde keuze')
            except:
                print('verkeerde input')
                sel=0
            else:
                if sel !=0:
                    product=ws.cell(row=prod_code+1,column=2).value
                    prijs=ws.cell(row=prod_code+1,column=4).value 
                    if ws.cell(row=prod_code+1,column=5).value is None:
                        korting=0
                    else:
                        korting=ws.cell(row=prod_code+1,column=5).value
                    aankoopnr=sub_sub_aankopen(klantnr,ws,wb,cat,prod_code,product,aantal,prijs,korting,nieuw) #naar procedure die de aankopen wegschrijft
                    nieuw=0 #vanaf nu is het niet meer een nieuw order 
                check=0
                while check==0:
                    try:
                        ja_nee=input('\nwilt u nog producten kiezen? (j/n): ')
                        if ja_nee in ("j","n"):
                            check=1
                    except:
                        print('verkeerde input')
                    else:
                        nog_prod=ja_nee #nodig om uit de buitenste loop te gaan als niet meer 'nog producten' (deze van nog_prod)
                        sel=1  #nodig om bij ja_nee=n, uit de binnenste loop te gaan (deze van sel)
    if aankoopnr != 0:
        sub_rekening(aankoopnr,klantnr)

def aankopen():
    klantnr=sub_klant()  #eerste stap: klantgegevens
    print('\nklantnummer is:',klantnr)
    try: #tweede stap: de producten in de winkel
        wb = load_workbook('Winkel.xlsx')
    except:
        producteninladen()
        wb=load_workbook('Winkel.xlsx')
    sub_aankopen(wb,klantnr) #derde stap: de effectieve aankopen
    wb.close()
    
def sub_winkeloverzicht(wb,dag,dag_nr,dagen):
    overzicht=()
    if dag_nr!=0:
        dagen=[dag] #we zetten in dit geval de lijst naar enkel één element: dag, om volgende for loop altijd te kunnen uitvoeren
        print('dag:',dag)
    else:
        print('voor alle dagen samen:')
        ws1=wb.create_sheet('Overzicht')
        ws1.cell(row=1,column=1).value="dag"
        ws1.cell(row=1,column=2).value="omzet"
    rij=1
    for dag_i in dagen: 
        ws=wb[dag_i]
        som=0
        for i in range(2,ws.max_row+1):
            som+=ws.cell(row=i,column=9).value
        aantal_aankopen=ws.cell(row=ws.max_row,column=1).value  #aantal aankopen = max aankoopnr
        gemiddeld=round(som/aantal_aankopen,2)
        overzicht=overzicht.__add__((dag_i,som,aantal_aankopen,gemiddeld))
        if dag_nr==0:
            ws1.cell(row=rij+1,column=1).value=dag_i
            ws1.cell(row=rij+1,column=2).value=som
        rij+=1
    tot_som=0
    tot_aantal_aankopen=0
    for teller in range(0,int(len(overzicht)/4)): # alles samen
        tot_som+=overzicht[4*teller+1]
        tot_aantal_aankopen+=overzicht[4*teller+2]
    print('totale som:',format(tot_som,'.2f'))
    print('totaal aantal aankopen:',tot_aantal_aankopen)
    print('gemiddeld bedrag per aankoop:',format(round(tot_som/tot_aantal_aankopen,2),'.2f'))
    if dag_nr==0:
        refObj=Reference(ws1,min_col=2,max_col=2,min_row=1,max_row=ws1.max_row)
        titles = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
        chartObj=BarChart3D()
        chartObj.title="omzet per dag"
        chartObj.add_data(data=refObj,titles_from_data=True)
        chartObj.set_categories(titles)
        ws1.add_chart(chartObj,'D3')
        wb.save('Inkopen.xlsx')
        print('excel sheet genereerd met grafiek: zie Inkopen.xlsx')
        wb.close()
    
def winkeloverzicht():
    dagen=[]
    try:
        wb=load_workbook('Inkopen.xlsx')
        teller=1
        print('er zijn aankopen gebeurd voor volgende dagen')
        for i in wb.sheetnames:
            if i != "Overzicht": #om de Overzicht worksheet over te slaan wanneer deze eerder gecreëerd werd bij vorige run
                ws=wb[i]
                dagen.append(i)
                print(teller,': datum: ',ws.title)
                teller+=1
    except:
        print('probleem openen bestand aankoopgegevens')
    else:
        check=0
        while check==0:
            try:
                dag_nr=int(input('\nvoor welke dag? (nummer) - of globaal overzicht (0 ingeven): '))
            except:
                print('verkeerde input')
            else:
                if (1<=dag_nr<teller):
                    check=1
                    dag=dagen[dag_nr-1]
                    sub_winkeloverzicht(wb,dag,dag_nr,dagen)
                elif dag_nr==0:
                    check=1
                    dag=dagen[0] #dummy gezet om te kunnen doorgeven als noodzakelijke parameter in volgende sub
                    try:
                        del wb["Overzicht"] #vorige versie van het Overzicht verwijderen
                    except:
                        sub_winkeloverzicht(wb,dag,dag_nr,dagen) 
                    else:
                        sub_winkeloverzicht(wb,dag,dag_nr,dagen)
                check1=0
                while check1==0:
                    try:
                        opnieuw=input('opnieuw? (j/n): ')
                        if opnieuw in ('j','n'):
                            check1=1
                    except:
                        print('verkeerde input')
                    else:
                        if opnieuw=='j':
                            check=0
                        elif opnieuw=="n":
                            check=1
        wb.close()

def selectie(actie):
    print('dit zijn de opties: \n(1): als klant: aankopen doen \n(2): als eigenaar: overzicht van de verkopen \n(3): als eigenaar: bijbestellingen plaatsen \n(4): als eigenaar: prijsaanpassingen \n(5): als eigenaar: promotiekortingen ingeven \n(6): als eigenaar: producten invoeren/verwijderen \n(7): stoppen')
    while actie==0:
        try:
            actie=int(input('wat wilt u doen?: '))
            if actie not in range(1,8):
                actie=0 
        except:
            print('verkeerde input')
            actie=0
    if actie==7:
        print('\ngestopt')
        repeat="n"
    elif actie==1:
        aankopen()
        repeat=""
    else:
        pwd=1
        while pwd!=13:
            try:
                pwd=int(input('wat is het paswoord?: (hint: 0 1 1 2 3 5 8 ..) (stoppen door 0 in te geven): '))
                if pwd==0:
                    break
                if pwd != 13:
                    print('verkeerd')
                    pwd=1
            except:
                print('verkeerde input')
                pwd=1
        if pwd==13:
            if actie==2:
                winkeloverzicht()
            elif actie in (3,4,5):
                bijbestellingen_prijsaanpassingen_promotiekortingen(actie) 
            elif actie==6:
                nieuwe_of_verwijderen_producten()
        repeat=""
    return repeat

def hoofdprogramma():
    repeat=""
    actie=0
    repeat=selectie(actie)
    while repeat=="":
        try:
            repeat=input('\nwil u het hoofdmenu opnieuw uitvoeren? (j/n): ')
            if repeat not in ("j","n"): 
                repeat=""
        except:
            print('verkeerde input\n')
            repeat="" 
    if repeat=="j":
        print('\n')
        hoofdprogramma()

hoofdprogramma()